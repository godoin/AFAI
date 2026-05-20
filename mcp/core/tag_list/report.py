from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from core.logger import logger
from core.tag_list.parser import ParseResult
from core.tag_list.validator import ValidationResult


# ── Shared style constants ─────────────────────────────────────────────────────

_FONT_NAME = "Arial"

# Header row
_HDR_FONT  = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
_HDR_FILL  = PatternFill("solid", start_color="1F3864")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Alternating row fill
_ALT_FILL  = PatternFill("solid", start_color="EEF2FF")

# Status cell fills
_FILL_CREATE   = PatternFill("solid", start_color="C6EFCE")   # green
_FILL_EXISTS   = PatternFill("solid", start_color="FFEB9C")   # amber
_FILL_AF_MISS  = PatternFill("solid", start_color="FFCCCC")   # red-light
_FILL_VIOLATION= PatternFill("solid", start_color="F2DCDB")   # salmon
_FILL_CREATED  = PatternFill("solid", start_color="C6EFCE")   # green
_FILL_FAILED   = PatternFill("solid", start_color="FFC7CE")   # red
_FILL_SKIPPED  = PatternFill("solid", start_color="FFEB9C")   # amber

_ACTION_FILLS = {
    "CREATE":                    _FILL_CREATE,
    "SKIP — already exists":     _FILL_EXISTS,
    "SKIP — AF element missing": _FILL_AF_MISS,
    "SKIP — naming violation":   _FILL_VIOLATION,
}

_STATUS_FILLS = {
    "CREATED":        _FILL_CREATED,
    "ALREADY EXISTED":_FILL_EXISTS,
    "FAILED":         _FILL_FAILED,
    "SKIPPED":        _FILL_SKIPPED,
}

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_DATA_FONT  = Font(name=_FONT_NAME, size=10)
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT_WRAP  = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Column definitions ─────────────────────────────────────────────────────────

# Pre-action report — Tags sheet
_PRE_ACTION_COLS = [
    ("Row #",                                   8,  "center"),
    ("Plant",                                   22, "center"),
    ("Unit/System",                             18, "center"),
    ("Source Tagname",                          30, "left"),
    ("Source Tag",                              28, "left"),
    ("Proposed New Tagname",                    34, "left"),
    ("Canary Tag Path",                         36, "left"),
    ("Description",                             38, "left"),
    ("eng_units",                               10, "center"),
    ("Date Added",                              14, "center"),
    ("Data Tag Naming for Checking (Remarks)",  40, "left"),
    ("Naming Valid",                            22, "center"),
    ("Exists in PI",                            14, "center"),
    ("Exists in AF",                            14, "center"),
    ("Proposed Action",                         30, "center"),
    ("BA Notes",                                30, "left"),
]

# Output report — Full results sheet (pre-action cols + final status cols)
_OUTPUT_EXTRA_COLS = [
    ("Final Status",        18, "center"),
    ("PI Tag Created",      16, "center"),
    ("AF Attribute Linked", 20, "center"),
    ("Live Data Received",  20, "center"),
    ("Error Detail",        40, "left"),
]

_OUTPUT_COLS = _PRE_ACTION_COLS + _OUTPUT_EXTRA_COLS


# ── Report generator ───────────────────────────────────────────────────────────

class TagListReportGenerator:
    """
    Generates two types of Excel report from the tag list workflow:

    1. Pre-action report (Gate 1)
       - Sheet 1: Summary  — counts by category
       - Sheet 2: Tags     — one row per tag with Proposed Action and BA Notes

    2. Output report (Phase 5)
       - Sheet 1: Session Summary — final counts + session metadata
       - Sheet 2: Full Results    — all pre-action columns + final status columns

    Both reports use the same formatting conventions as the client tag list
    so the BA can compare them side by side.

    Usage:
        gen = TagListReportGenerator()

        # Gate 1
        path = gen.pre_action_report(validation_result, output_dir=".")

        # Phase 5 — pass the same validation_result with rows updated
        path = gen.output_report(
            validation_result,
            af_database_path="PI-SYSTEM/GoogleManualLogger",
            output_dir="."
        )
    """

    def __init__(self):
        pass

    # ── Public: pre-action report ─────────────────────────────────────────────

    def pre_action_report(
        self,
        validation_result: ValidationResult,
        output_dir: str = ".",
        filename: str = None,
    ) -> str:
        """
        Build and save the Gate 1 pre-action report.

        Returns the absolute path of the saved file.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname     = filename or f"PreAction_Report_{timestamp}.xlsx"
        out_path  = Path(output_dir) / fname

        wb = Workbook()

        self._build_summary_sheet(wb, validation_result)
        self._build_tags_sheet(wb, validation_result)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(out_path))
        logger.info(f"Pre-action report saved: {out_path}", exc_info=False)
        return str(out_path)

    # ── Public: output report ─────────────────────────────────────────────────

    def output_report(
        self,
        validation_result: ValidationResult,
        af_database_path: str = "",
        output_dir: str = ".",
        filename: str = None,
    ) -> str:
        """
        Build and save the Phase 5 session output report.

        Rows in validation_result should have their Final Status fields
        populated by Phase 4 before calling this method.

        Returns the absolute path of the saved file.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname     = filename or f"Output_Report_{timestamp}.xlsx"
        out_path  = Path(output_dir) / fname

        wb = Workbook()

        self._build_session_summary_sheet(wb, validation_result, af_database_path)
        self._build_full_results_sheet(wb, validation_result)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(out_path))
        logger.info(f"Output report saved: {out_path}", exc_info=False)
        return str(out_path)

    # ── Private: pre-action sheets ────────────────────────────────────────────

    def _build_summary_sheet(
        self,
        wb: Workbook,
        vr: ValidationResult,
    ) -> None:
        ws = wb.create_sheet("Summary")
        s  = vr.summary()

        # Title
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value     = "Pre-Action Report — Summary"
        title_cell.font      = Font(name=_FONT_NAME, bold=True, size=12, color="FFFFFF")
        title_cell.fill      = _HDR_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Generated timestamp
        ws.merge_cells("A2:B2")
        ts_cell = ws["A2"]
        ts_cell.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ts_cell.font      = Font(name=_FONT_NAME, size=9, italic=True, color="595959")
        ts_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        # Spacer
        ws.row_dimensions[3].height = 6

        # Summary rows
        summary_rows = [
            ("Total rows in file",      s.get("Total rows in file",      0), None),
            ("Rows passing validation", s.get("Rows passing validation", 0), None),
            ("",                        "",                                   None),
            ("Tags to create",          s.get("Tags to create",          0), _FILL_CREATE),
            ("Tags already existing",   s.get("Tags already existing",   0), _FILL_EXISTS),
            ("AF elements missing",     s.get("AF elements missing",     0), _FILL_AF_MISS),
            ("Naming violations",       s.get("Naming violations",       0), _FILL_VIOLATION),
            ("",                        "",                                   None),
            ("Rows skipped (total)",    s.get("Rows skipped (total)",    0), None),
        ]

        label_font = Font(name=_FONT_NAME, size=10)
        value_font = Font(name=_FONT_NAME, size=10, bold=True)

        for i, (label, value, fill) in enumerate(summary_rows, start=4):
            lc = ws.cell(row=i, column=1, value=label)
            vc = ws.cell(row=i, column=2, value=value)
            lc.font      = label_font
            vc.font      = value_font
            lc.alignment = _LEFT_WRAP
            vc.alignment = _CENTER
            if fill and label:
                lc.fill = fill
                vc.fill = fill
            if label:
                lc.border = _BORDER
                vc.border = _BORDER

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14

    def _build_tags_sheet(
        self,
        wb: Workbook,
        vr: ValidationResult,
    ) -> None:
        ws = wb.create_sheet("Tags")
        headers = [c[0] for c in _PRE_ACTION_COLS]
        widths  = [c[1] for c in _PRE_ACTION_COLS]
        aligns  = [c[2] for c in _PRE_ACTION_COLS]

        self._write_header_row(ws, headers)

        for row_idx, tag_row in enumerate(vr.all_rows, start=2):
            d    = tag_row.to_dict()
            fill = _ALT_FILL if row_idx % 2 == 0 else None
            action_fill = _ACTION_FILLS.get(tag_row.proposed_action)

            for col_idx, header in enumerate(headers, start=1):
                value = d.get(header, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = _DATA_FONT
                cell.border = _BORDER

                # Proposed Action column gets status colour
                if header == "Proposed Action" and action_fill:
                    cell.fill = action_fill
                elif fill:
                    cell.fill = fill

                align_str = aligns[col_idx - 1]
                cell.alignment = _CENTER if align_str == "center" else _LEFT_WRAP

        self._apply_column_widths(ws, widths)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}1"
        )

    # ── Private: output sheets ────────────────────────────────────────────────

    def _build_session_summary_sheet(
        self,
        wb: Workbook,
        vr: ValidationResult,
        af_database_path: str,
    ) -> None:
        ws = wb.create_sheet("Session Summary")

        # Title
        ws.merge_cells("A1:B1")
        t = ws["A1"]
        t.value     = "Output Report — Session Summary"
        t.font      = Font(name=_FONT_NAME, bold=True, size=12, color="FFFFFF")
        t.fill      = _HDR_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:B2")
        ts = ws["A2"]
        ts.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ts.font      = Font(name=_FONT_NAME, size=9, italic=True, color="595959")
        ts.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6

        # Compute output counts from row fields
        created     = sum(1 for r in vr.all_rows if r.final_status == "CREATED")
        existed     = sum(1 for r in vr.all_rows if r.final_status == "ALREADY EXISTED")
        failed      = sum(1 for r in vr.all_rows if r.final_status == "FAILED")
        skipped     = sum(1 for r in vr.all_rows if r.final_status == "SKIPPED")
        linked      = sum(1 for r in vr.all_rows if r.af_attribute_linked == "Yes")
        no_data     = sum(1 for r in vr.all_rows if r.live_data_received == "No")

        summary_rows = [
            ("Session date",                   datetime.now().strftime("%Y-%m-%d"), None),
            ("AF database",                    af_database_path,                   None),
            ("",                               "",                                  None),
            ("Tags created successfully",      created,  _FILL_CREATED),
            ("Tags already existed (skipped)", existed,  _FILL_EXISTS),
            ("Tags failed",                    failed,   _FILL_FAILED),
            ("Tags skipped (naming violation)",
             len(vr.naming_violations),        _FILL_VIOLATION),
            ("Tags skipped (AF element missing)",
             len(vr.af_missing),               _FILL_AF_MISS),
            ("",                               "",                                  None),
            ("Attributes linked successfully", linked,   _FILL_CREATE),
            ("Attributes with no live data",   no_data,  _FILL_EXISTS),
        ]

        lf = Font(name=_FONT_NAME, size=10)
        vf = Font(name=_FONT_NAME, size=10, bold=True)

        for i, (label, value, fill) in enumerate(summary_rows, start=4):
            lc = ws.cell(row=i, column=1, value=label)
            vc = ws.cell(row=i, column=2, value=value)
            lc.font      = lf
            vc.font      = vf
            lc.alignment = _LEFT_WRAP
            vc.alignment = _CENTER
            if fill and label:
                lc.fill = fill
                vc.fill = fill
            if label:
                lc.border = _BORDER
                vc.border = _BORDER

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 28

    def _build_full_results_sheet(
        self,
        wb: Workbook,
        vr: ValidationResult,
    ) -> None:
        ws = wb.create_sheet("Full Results")
        headers = [c[0] for c in _OUTPUT_COLS]
        widths  = [c[1] for c in _OUTPUT_COLS]
        aligns  = [c[2] for c in _OUTPUT_COLS]

        self._write_header_row(ws, headers)

        for row_idx, tag_row in enumerate(vr.all_rows, start=2):
            d    = tag_row.to_dict()
            fill = _ALT_FILL if row_idx % 2 == 0 else None

            # Pull Phase 4 fields
            final_status    = tag_row.final_status
            pi_created      = tag_row.pi_tag_created
            af_linked       = tag_row.af_attribute_linked
            live_data       = tag_row.live_data_received
            error_detail    = tag_row.error_detail

            extra = {
                "Final Status":        final_status,
                "PI Tag Created":      pi_created,
                "AF Attribute Linked": af_linked,
                "Live Data Received":  live_data,
                "Error Detail":        error_detail,
            }
            d.update(extra)

            status_fill  = _STATUS_FILLS.get(final_status)
            action_fill  = _ACTION_FILLS.get(tag_row.proposed_action)

            for col_idx, header in enumerate(headers, start=1):
                value = d.get(header, "")
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = _DATA_FONT
                cell.border = _BORDER

                if header == "Final Status" and status_fill:
                    cell.fill = status_fill
                elif header == "Proposed Action" and action_fill:
                    cell.fill = action_fill
                elif fill:
                    cell.fill = fill

                align_str = aligns[col_idx - 1]
                cell.alignment = _CENTER if align_str == "center" else _LEFT_WRAP

        self._apply_column_widths(ws, widths)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}1"
        )

    # ── Private: shared helpers ───────────────────────────────────────────────

    @staticmethod
    def _write_header_row(ws, headers: list) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = _HDR_FONT
            cell.fill      = _HDR_FILL
            cell.alignment = _HDR_ALIGN
            cell.border    = _BORDER
        ws.row_dimensions[1].height = 36

    @staticmethod
    def _apply_column_widths(ws, widths: list) -> None:
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width