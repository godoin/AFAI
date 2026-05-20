"""
test_report.py — unit tests for core.tag_list.report.TagListReportGenerator

Run:  pytest tests/test_report.py -v
"""

import os
import pytest
from openpyxl import load_workbook

from core.tag_list.parser import TagListParser
from core.tag_list.validator import TagListValidator
from core.tag_list.report import TagListReportGenerator, _PRE_ACTION_COLS, _OUTPUT_COLS
from tests.conftest import MockElements, MockPoints, MockPISystem


@pytest.fixture
def parser():
    return TagListParser()


@pytest.fixture
def generator():
    return TagListReportGenerator()


def make_validation_result(parser, file_path, af_db, existing_tags=None, existing_plants=None):
    """Helper: parse + validate in one call."""
    pr = parser.parse(file_path)
    vr = TagListValidator(
        elements=MockElements(existing_plants or {"PlantA", "PlantB"}),
        points=MockPoints(existing_tags or set()),
        pi_system=MockPISystem()
    ).validate(pr, af_db)
    return vr


# ── Pre-action report ──────────────────────────────────────────────────────────

class TestPreActionReport:

    def test_creates_file(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        assert os.path.exists(path)

    def test_has_summary_and_tags_sheets(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Tags"    in wb.sheetnames
        assert "Sheet"   not in wb.sheetnames

    def test_no_default_empty_sheet(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        assert set(wb.sheetnames) == {"Summary", "Tags"}

    def test_summary_counts_correct(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        existing = {"PlantA_Unit1_Flow_Rate"}   # 1 already exists
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path, existing_tags=existing)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Summary"]

        pairs = {r[0].value: r[1].value for r in ws.iter_rows(min_row=4) if r[0].value}
        assert pairs["Total rows in file"]      == 5
        assert pairs["Tags to create"]          == 4
        assert pairs["Tags already existing"]   == 1
        assert pairs["AF elements missing"]     == 0
        assert pairs["Naming violations"]       == 0

    def test_tags_sheet_headers_match_spec(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr     = make_validation_result(parser, valid_taglist_file, af_database_path)
        path   = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb     = load_workbook(path)
        ws     = wb["Tags"]
        expected = [c[0] for c in _PRE_ACTION_COLS]
        actual   = [ws.cell(1, i + 1).value for i in range(len(expected))]
        assert actual == expected

    def test_tags_sheet_row_count(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Tags"]
        assert ws.max_row - 1 == 5    # 5 data rows + 1 header

    def test_proposed_action_values(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        existing = {"PlantA_Unit1_Flow_Rate"}
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path, existing_tags=existing)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Tags"]

        headers   = [c[0] for c in _PRE_ACTION_COLS]
        action_col = headers.index("Proposed Action") + 1
        actions   = [ws.cell(r, action_col).value for r in range(2, ws.max_row + 1)]

        assert actions.count("CREATE")               == 4
        assert actions.count("SKIP — already exists") == 1

    def test_naming_violations_in_tags_sheet(self, parser, violation_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, violation_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Tags"]

        headers    = [c[0] for c in _PRE_ACTION_COLS]
        action_col = headers.index("Proposed Action") + 1
        actions    = [ws.cell(r, action_col).value for r in range(2, ws.max_row + 1)]

        assert actions.count("SKIP — naming violation") == 4
        assert actions.count("CREATE")                  == 1

    def test_freeze_pane_and_autofilter(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Tags"]
        assert ws.freeze_panes       == "A2"
        assert ws.auto_filter.ref   is not None

    def test_custom_filename(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path), filename="my_report.xlsx")
        assert path.endswith("my_report.xlsx")
        assert os.path.exists(path)

    def test_ba_notes_column_is_blank(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        path = generator.pre_action_report(vr, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Tags"]

        headers  = [c[0] for c in _PRE_ACTION_COLS]
        ba_col   = headers.index("BA Notes") + 1
        ba_vals  = [ws.cell(r, ba_col).value for r in range(2, ws.max_row + 1)]
        assert all(v in (None, "") for v in ba_vals)


# ── Output report ──────────────────────────────────────────────────────────────

class TestOutputReport:

    def _annotate_rows(self, vr, create_status="CREATED", exist_status="ALREADY EXISTED"):
        """Simulate Phase 4 by annotating rows with final status fields."""
        for row in vr.to_create:
            row.final_status        = create_status
            row.pi_tag_created      = "Yes" if create_status == "CREATED" else "No"
            row.af_attribute_linked = "Yes" if create_status == "CREATED" else "No"
            row.live_data_received  = "Yes" if create_status == "CREATED" else "Not checked"
        for row in vr.already_exist:
            row.final_status        = exist_status
            row.pi_tag_created      = "No"
            row.af_attribute_linked = "No"
            row.live_data_received  = "Not checked"

    def test_creates_file(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        assert os.path.exists(path)

    def test_has_session_summary_and_full_results_sheets(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb = load_workbook(path)
        assert "Session Summary" in wb.sheetnames
        assert "Full Results"    in wb.sheetnames
        assert "Sheet"           not in wb.sheetnames

    def test_full_results_headers_match_spec(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Full Results"]

        expected = [c[0] for c in _OUTPUT_COLS]
        actual   = [ws.cell(1, i + 1).value for i in range(len(expected))]
        assert actual == expected

    def test_full_results_row_count(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Full Results"]
        assert ws.max_row - 1 == 5

    def test_final_status_values(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        existing = {"PlantA_Unit1_Flow_Rate"}
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path, existing_tags=existing)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Full Results"]

        headers    = [c[0] for c in _OUTPUT_COLS]
        status_col = headers.index("Final Status") + 1
        statuses   = [ws.cell(r, status_col).value for r in range(2, ws.max_row + 1)]

        assert statuses.count("CREATED")         == 4
        assert statuses.count("ALREADY EXISTED") == 1

    def test_session_summary_counts(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        existing = {"PlantA_Unit1_Flow_Rate"}
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path, existing_tags=existing)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Session Summary"]

        pairs = {r[0].value: r[1].value for r in ws.iter_rows(min_row=4) if r[0].value}
        assert pairs["Tags created successfully"]      == 4
        assert pairs["Tags already existed (skipped)"] == 1
        assert pairs["Tags failed"]                    == 0
        assert pairs["Attributes linked successfully"] == 4

    def test_failed_rows_captured(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr, create_status="FAILED")
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Session Summary"]

        pairs = {r[0].value: r[1].value for r in ws.iter_rows(min_row=4) if r[0].value}
        assert pairs["Tags failed"] == 5

    def test_rows_in_file_order(self, parser, valid_taglist_file, af_database_path, generator, tmp_path):
        vr   = make_validation_result(parser, valid_taglist_file, af_database_path)
        self._annotate_rows(vr)
        path = generator.output_report(vr, af_database_path, output_dir=str(tmp_path))
        wb   = load_workbook(path)
        ws   = wb["Full Results"]

        headers  = [c[0] for c in _OUTPUT_COLS]
        row_col  = headers.index("Row #") + 1
        row_nums = [ws.cell(r, row_col).value for r in range(2, ws.max_row + 1)]
        assert row_nums == sorted(row_nums)